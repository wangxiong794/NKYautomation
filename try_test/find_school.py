# coding=utf-8
import os
import time

from selenium.webdriver.firefox.options import Options

from service.logger import Log
import openpyxl
import xlrd
from selenium import webdriver
from selenium.webdriver.common.by import By


def read(row):
    book = xlrd.open_workbook(r"F:\workspace\公司项目\双高招投标收集\197所国双高名单.xls")
    sheet = book.sheet_by_index(0)
    # print(sheet.cell_value(1, 1))
    return [sheet.cell_value(row, 1), sheet.cell_value(row, 2)]


# def selenium(company_name="北京电子科技职业学院"):
#     chrome_options = Options()
#     chrome_options.add_argument('--headless')
#     chrome_options.add_argument('--disable-gpu')
#     driver = webdriver.Chrome(chrome_options=chrome_options, executable_path='../chromedriver.exe')
#     driver = webdriver.Chrome(executable_path='../chromedriver.exe')
#     driver.get("https://www.baidu.com")
#     driver.find_element(By.XPATH, '//*[@id="kw"]').send_keys(company_name)
#     driver.find_element(By.XPATH, '//input[@id="su"]').click()
#     time.sleep(1)
#     print(company_name)
#     driver.find_element(By.XPATH, "//div[@id='content_left']/div[1]/div/div[1]/h3/a[1]").click()
#     time.sleep(1)
#     w = driver.window_handles
#     w1 = driver.current_window_handle
#     w2 = w.index(w1) + 1
#     driver.switch_to.window(w[w2])
#     url = driver.current_url
#     driver.quit()
#     return str(url)


def sg(keyword="中期自评报告"):
    log = Log()
    for i in range(1, 198):
        cname = read(i)
        name = cname[0]
        url = str(cname[1])
        log.info("正在查找："+str(name))
        # chrome_options = Options()
        # chrome_options.add_argument('--headless')
        # chrome_options.add_argument('--disable-gpu')
        # driver = webdriver.Chrome(chrome_options=chrome_options, executable_path='../chromedriver.exe')
        firefox_options = Options()
        firefox_options.add_argument('--headless')
        firefox_options.add_argument('--disable-gpu')
        driver = webdriver.Firefox(options=firefox_options, executable_path='../geckodriver.exe')
        # driver = webdriver.Chrome(executable_path='../chromedriver.exe')
        try:
            driver.get(url)
            time.sleep(1)
            a = driver.execute_script("return document.documentElement.outerHTML")
        # print(a)
            if keyword in str(a):
                # print(str(name)+":"+str(url))
                log.info(str(name)+"单位存在报告：《"+str(url)+"》")
                with open(r'log.txt','a+',encoding='utf-8') as f:
                    f.write(str(time.strftime('%Y%m%d-%H:%M:%S: '))+str(name)+"单位存在报告：《"+str(url)+"》\n")
                    f.close()
                driver.quit()
            else:
                log.warning(str(name) + "单位不存在")
                driver.quit()
        except:
            driver.quit()
            log.warning(str(name)+"单位打开失败，域名是"+str(url))
        os.system('taskkill /F /iM firefox.exe')



# def write():
#     # book = openpyxl.Workbook()
#     # sheet = book.create_sheet('url')
#     for i in range(1, 198):
#         cname = read(i)
#         url = selenium(cname)
#         print(cname + ":" + url)
#         # sheet.cell(i, 1).value(cname)
#         # sheet.cell(i, 2).value(url)
#         # print(cname + ":" + url)
#     # book.save('test.xlsx')


if __name__ == "__main__":
    # a=read(1)
    # print(a[0])
    a = input("请输入关键字(默认中期自评报告)：")
    if a is None:
        sg()
    else:
        sg(str(a))
